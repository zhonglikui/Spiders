# encoding:utf-8
import os
import time

import requests
import base64
from openpyxl import Workbook

api_key = "GXYzfmt42Gb7DoTW4j9QgN5c"
secret_key = "wGgn9dCiI8zBBrq5nVtt2EfmU66enntV"
tokenUrl = "https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id={}&client_secret={}"
#ocrUrl = "https://aip.baidubce.com/rest/2.0/ocr/v1/general_basic?access_token={}"
ocrUrl = "https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic?access_token={}"
list = []


def getToken():
    try:
        response = requests.get(tokenUrl.format(api_key, secret_key))
        if response:
            print("getToken:{}".format(response.json()))
            return response.json().get("access_token")
    except Exception as e:
        print(e)


def getResule(filePath, token):
    try:
        f = open('{}'.format(filePath), 'rb')
        img = base64.b64encode(f.read())
        params = {"image": img}
        request_url = ocrUrl.format(token)
        headers = {'content-type': 'application/x-www-form-urlencoded'}
        response = requests.post(request_url, data=params, headers=headers)
        if response:
            return response.json()
    except Exception as e:
        print(e)


def getImages(folderPath):
    if os.path.exists(folderPath) & os.path.isdir(folderPath):
        for root, dirs, files in os.walk(folderPath):
            for d in dirs:
                getImages(d)
            for f in files:
                type = f.split(".")[-1]
                if type == "jpg" or type == "png":
                    absPath = os.path.abspath(f)
                    print("filePath:{}".format(absPath))
                    if absPath not in list:
                        list.append(absPath)


if __name__ == "__main__":
    folderPath = os.getcwd()
    getImages(folderPath)
    token = getToken()
    wb = Workbook()
    ws = wb.active
    c1 = 0
    c2 = 2
    c3 = 4
    c4 = 6
    c5 = 8
    c6 = 10
    c7 = 12
    c8 = 14
    ws.cell(1, 1, ).value = "姓名"
    ws.cell(1, 2).value = "别名"
    ws.cell(1, 3).value = "电子邮件"
    ws.cell(1, 4).value = "电话"
    ws.cell(1, 5).value = "移动电话"
    ws.cell(1, 6).value = "职务"
    ws.cell(1, 7).value = "部门"
    ws.cell(1, 8).value = "公司"
    for index, file in enumerate(list):
        print("开始请求第{}/{}条：{}".format(index, len(list), file))
        result = getResule(file, token)
        # print("result: {}".format(result))
        # id = result.get("log_id")
        # num = result.get("words_result_num")
        words_result = result.get("words_result")
        # print("结果：id:{} num:{}".format(id, num))
        # 3、姓名 6、部门  8.工号  10 部门 13电话
        row = index + 2
        for indexColum, word in enumerate(words_result):
            print("{} : {}".format(indexColum, word.get("words")))
            continue
            value = word.get("words")
            if indexColum == c1:
                ws.cell(row, 1).value = value
            elif indexColum == c2:
                ws.cell(row, 2).value = value
            elif indexColum == c3:
                ws.cell(row, 3).value = value
            elif indexColum == c4:
                ws.cell(row, 4).value = value
            elif indexColum == c5:
                ws.cell(row, 5).value = value
            elif indexColum == c6:
                ws.cell(row, 6).value = value
            elif indexColum == c7:
                ws.cell(row, 7).value = value
            elif indexColum == c8:
                ws.cell(row, 8).value = value

        wb.save("地产开发-中电事业部.xlsx")
        print("解析完一张---------------------------------------------------------------------")
        time.sleep(1)
